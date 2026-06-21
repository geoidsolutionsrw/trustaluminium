from django.shortcuts import render, redirect, get_object_or_404
from product.models import Product, Supplier, Customer
from sale.models import Sale, SaleItem
from quotation.models import Quotation, QuotationItem
from django.contrib import messages
from django.contrib.auth.models import User
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
import pandas as pd
import os
from datetime import datetime, timedelta
from django.conf import settings
from django.db.models import (
    Sum, Count, F, Value, IntegerField, DecimalField,
    CharField, Case, When, Max, Min, ExpressionWrapper
)
from django.db.models.functions import (
    ExtractMonth, ExtractDay, TruncMonth,
    TruncWeek, TruncYear, Coalesce, Round
)
from django.utils import timezone
from calendar import month_abbr
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from decimal import Decimal
import json
import traceback
from dash.models import QuotationNotification


# ─────────────────────────────────────────────────────────────────────────────
# LANDING PAGE
# ─────────────────────────────────────────────────────────────────────────────

def homepage(request):
    """Landing page with live product catalog filtered by brand."""
    landing_products = Product.objects.all().order_by('product_name')
    all_products     = Product.objects.all().order_by('product_name')
    product_count    = Product.objects.count()
    customer_count   = Customer.objects.count()

    product_brands = (
        Product.objects
        .exclude(brand__isnull=True)
        .exclude(brand__exact='')
        .values_list('brand', flat=True)
        .distinct()
        .order_by('brand')
    )

    context = {
        'landing_products':  landing_products,
        'all_products':      all_products,
        'product_brands':    product_brands,
        'product_count':     product_count,
        'customer_count':    customer_count,
        'has_more_products': False,
    }
    return render(request, 'home/index.html', context)


def contactus(request):
    return render(request, 'home/contact_us.html')


# ─────────────────────────────────────────────────────────────────────────────
# PUBLIC QUOTATION REQUEST (from landing page)
# ─────────────────────────────────────────────────────────────────────────────

def public_quotation_request(request):
    """
    Handles quotation form from the landing page.
    Creates/finds a Customer, creates a PENDING Quotation,
    creates a QuotationNotification for staff, returns JSON.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'})

    customer_name = request.POST.get('customer_name', '').strip()
    phone_number  = request.POST.get('phone_number',  '').strip()
    email_address = request.POST.get('email_address', '').strip()
    company       = request.POST.get('company',       '').strip()
    notes         = request.POST.get('notes',         '').strip()
    product_ids   = request.POST.getlist('product_id[]')
    quantities    = request.POST.getlist('quantity[]')

    if not customer_name:
        return JsonResponse({'success': False, 'error': 'Full name is required.'})
    if not phone_number:
        return JsonResponse({'success': False, 'error': 'Phone number is required.'})
    if not product_ids or not any(pid.strip() for pid in product_ids):
        return JsonResponse({'success': False, 'error': 'Please select at least one product.'})

    try:
        # Get or create customer
        customer = Customer.objects.filter(phone_number=phone_number).first()
        if not customer:
            customer = Customer.objects.create(
                customer_name = customer_name,
                phone_number  = phone_number,
                email_address = email_address,
                address       = company,
                customer_type = 'INDIVIDUAL',
            )

        # Get system creator (first superuser/staff)
        creator = (
            User.objects.filter(is_superuser=True).first()
            or User.objects.filter(is_staff=True).first()
            or User.objects.first()
        )
        if not creator:
            return JsonResponse({'success': False, 'error': 'System configuration error. Please contact us directly.'})

        # Build quotation items
        items_data   = []
        total_amount = Decimal('0.00')

        for pid, qty in zip(product_ids, quantities):
            pid = pid.strip()
            if not pid:
                continue
            try:
                product  = Product.objects.get(productid=pid)
                quantity = max(int(qty), 1) if qty else 1
                price    = Decimal(str(
                    getattr(product, 'selling_price', None)
                    or getattr(product, 'price', None)
                    or 0
                ))
                subtotal     = price * quantity
                total_amount += subtotal
                items_data.append({
                    'product':  product,
                    'quantity': quantity,
                    'price':    price,
                    'subtotal': subtotal,
                })
            except (Product.DoesNotExist, ValueError):
                continue

        if not items_data:
            return JsonResponse({'success': False, 'error': 'No valid products selected.'})

        # Create Quotation
        quotation = Quotation.objects.create(
            customer         = customer,
            date             = timezone.now(),
            total_amount     = total_amount,
            quotation_status = 'PENDING',
            creator          = creator,
        )

        # Create QuotationItems
        for item in items_data:
            QuotationItem.objects.create(
                quotation = quotation,
                product   = item['product'],
                quantity  = item['quantity'],
                price     = item['price'],
                tax       = Decimal('0.00'),
                subtotal  = item['subtotal'],
            )

        # ── Create staff notification ──────────────────────────────────
        QuotationNotification.objects.create(
            quotation_id  = quotation.quo_id,
            customer_name = customer_name,
            phone_number  = phone_number,
            product_count = len(items_data),
        )

        return JsonResponse({
            'success':      True,
            'quotation_id': quotation.quo_id,
            'message':      f'Quotation #{quotation.quo_id} created successfully.',
        })

    except Exception:
        print(f"Quotation creation error:\n{traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'error':   'An error occurred. Please try again or contact us directly.'
        })


# ─────────────────────────────────────────────────────────────────────────────
# NOTIFICATION API ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────

def get_notifications(request):
    """
    Polled every 30s — returns unread notifications for this user.
    Returns empty list if not authenticated (no redirect).
    """
    if not request.user.is_authenticated:
        return JsonResponse({'notifications': []})

    unread = QuotationNotification.objects.exclude(
        read_by=request.user
    ).order_by('-created_at')[:10]

    notifications = [{
        'id':            n.id,
        'quotation_id':  n.quotation_id,
        'customer_name': n.customer_name,
        'phone_number':  n.phone_number,
        'product_count': n.product_count,
        'created_at':    n.created_at.strftime('%d %b %Y, %H:%M'),
    } for n in unread]

    return JsonResponse({'notifications': notifications})


@login_required
def mark_notification_read(request, notification_id):
    """Mark one notification as read for this user."""
    if request.method == 'POST':
        try:
            notif = QuotationNotification.objects.get(id=notification_id)
            notif.read_by.add(request.user)
            return JsonResponse({'success': True})
        except QuotationNotification.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Not found'})
    return JsonResponse({'success': False})


# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url="/login")
def home(request):
    from_date    = request.GET.get('from_date')
    to_date      = request.GET.get('to_date')
    period       = request.GET.get('period', '')
    chart_period = request.GET.get('chart_period', 'monthly')

    try:
        from_date = datetime.strptime(from_date, '%d-%m-%Y') if from_date else None
        to_date   = datetime.strptime(to_date,   '%d-%m-%Y') if to_date   else None
    except ValueError:
        from_date = to_date = None

    today = timezone.now()

    if period == 'today':
        from_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        to_date   = today
    elif period == 'week':
        from_date = today - timedelta(days=today.weekday())
        to_date   = today
    elif period == 'month':
        from_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        to_date   = today
    elif period == 'year':
        from_date = today.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        to_date   = today

    products   = Product.objects.values('productid', 'product_name').order_by('product_name')
    product_id = request.GET.get('product_id', '')
    is_admin   = request.user.groups.filter(name='Administrators').exists()
    selected_user_id = request.GET.get('user_id') if is_admin else None

    try:
        product_id = int(product_id) if product_id else None
    except ValueError:
        product_id = None

    sales_query = Sale.objects.all() if is_admin else Sale.objects.filter(creator=request.user)
    if is_admin and selected_user_id:
        sales_query = sales_query.filter(creator_id=selected_user_id)

    sale_query = sales_query.values(
        'customer__customer_name', 'customer__tin',
        'payment_status', 'date', 'total_amount',
        'sale_id', 'creator__username',
    ).annotate(
        paid_amount=Coalesce(Sum('payments__paid_amount'), Value(0), output_field=IntegerField()),
        balance=ExpressionWrapper(
            F('total_amount') - Coalesce(Sum('payments__paid_amount'), Value(0), output_field=IntegerField()),
            output_field=IntegerField()
        ),
        payment_completion=Case(
            When(
                total_amount=Coalesce(Sum('payments__paid_amount'), Value(0), output_field=DecimalField()),
                then=Value('Complete')
            ),
            default=Value('Incomplete'),
            output_field=CharField(),
        )
    ).order_by('-date')

    if from_date:
        sale_query = sale_query.filter(date__gte=from_date)
    if to_date:
        sale_query = sale_query.filter(date__lte=to_date)
    if product_id:
        sale_query = sale_query.filter(items__product__productid=product_id)

    total_paid        = sale_query.aggregate(total=Sum('payments__paid_amount'))['total'] or 0
    total_to_be_paid  = sale_query.aggregate(total=Sum('total_amount'))['total'] or 0
    remaining_balance = total_to_be_paid - total_paid
    total_items       = sale_query.aggregate(total=Sum('items__quantity'))['total'] or 0

    customer_count     = Customer.objects.count()
    sale_count         = Sale.objects.count()
    quotation_count    = Quotation.objects.count()
    registered_product = Product.objects.count()

    current_year  = today.year
    current_month = today.month

    this_month_sales = Sale.objects.filter(
        date__year=current_year, date__month=current_month
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    last_month_date  = today.replace(day=1) - timedelta(days=1)
    last_month_sales = Sale.objects.filter(
        date__year=last_month_date.year, date__month=last_month_date.month
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    sales_growth = (
        round(((this_month_sales - last_month_sales) / last_month_sales) * 100, 1)
        if last_month_sales > 0
        else (100 if this_month_sales > 0 else 0)
    )

    monthly_labels = list(month_abbr)[1:]
    monthly_totals = [0] * 12

    if chart_period == 'weekly':
        weekly_sales = sale_query.filter(date__year=current_year).annotate(
            week=TruncWeek('date')
        ).values('week').annotate(total=Sum('total_amount')).order_by('week')
        chart_labels = [f"W{s['week'].isocalendar()[1]}" for s in weekly_sales]
        chart_data   = [float(s['total'] or 0) for s in weekly_sales]
        chart_title  = 'Weekly Sales'
    elif chart_period == 'yearly':
        yearly_sales = Sale.objects.annotate(year=TruncYear('date')).values('year').annotate(
            total=Sum('total_amount')
        ).order_by('year')
        chart_labels = [str(s['year'].year) for s in yearly_sales]
        chart_data   = [float(s['total'] or 0) for s in yearly_sales]
        chart_title  = 'Yearly Sales'
    else:
        monthly_sales = sale_query.filter(date__year=current_year).annotate(
            month=TruncMonth('date')
        ).values('month').annotate(total=Sum('total_amount')).order_by('month')
        for s in monthly_sales:
            monthly_totals[s['month'].month - 1] = float(s['total'] or 0)
        chart_labels = monthly_labels
        chart_data   = monthly_totals
        chart_title  = 'Monthly Sales'

    payment_data = sale_query.filter(date__year=current_year).annotate(
        month=TruncMonth('date')
    ).values('month', 'payment_status').annotate(
        total=Sum('total_amount')
    ).order_by('month', 'payment_status')

    months, full_amounts, partial_amounts, pending_amounts = [], [], [], []
    current_month_data = None
    month_data = {'FULL': 0, 'PARTIAL': 0, 'PENDING': 0}

    for payment in payment_data:
        payment_month = payment['month']
        if current_month_data != payment_month:
            if current_month_data is not None:
                full_amounts.append(float(month_data['FULL']))
                partial_amounts.append(float(month_data['PARTIAL']))
                pending_amounts.append(float(month_data['PENDING']))
            month_data = {'FULL': 0, 'PARTIAL': 0, 'PENDING': 0}
            months.append(payment_month.strftime('%b'))
            current_month_data = payment_month
        if payment['payment_status'] and payment['total']:
            month_data[payment['payment_status']] = float(payment['total'])

    if current_month_data is not None:
        full_amounts.append(float(month_data['FULL']))
        partial_amounts.append(float(month_data['PARTIAL']))
        pending_amounts.append(float(month_data['PENDING']))

    top_products_chart = (
        SaleItem.objects.values('product__product_name')
        .annotate(total_revenue=Sum('subtotal'))
        .order_by('-total_revenue')[:8]
    )
    product_chart_labels = [p['product__product_name'] for p in top_products_chart]
    product_chart_data   = [float(p['total_revenue'] or 0) for p in top_products_chart]

    top_selling_products = (
        SaleItem.objects.values(
            'product__product_name', 'product__image', 'product_id'
        ).annotate(total=Sum('subtotal'), quantity=Sum('quantity')).order_by('-quantity')[:10]
    )
    processed_items = [{
        'id':        item['product_id'],
        'name':      item['product__product_name'],
        'total':     item['total'],
        'quantity':  item['quantity'],
        'image_url': (
            f"{settings.MEDIA_URL}{item['product__image']}"
            if item['product__image']
            else f"{settings.STATIC_URL}img/noimage.png"
        )
    } for item in top_selling_products]

    top_customers_qs = (
        SaleItem.objects.values(
            'sale__customer__customer_name',
            'sale__customer__cust_id',
            'sale__customer__email_address',
        ).annotate(
            total_amount=Coalesce(Sum('subtotal'), Value(0), output_field=DecimalField()),
            total_items =Coalesce(Sum('quantity'), Value(0), output_field=IntegerField()),
            total_orders=Count('sale', distinct=True),
        ).order_by('-total_amount')[:5]
    )

    users_with_sales = (
        User.objects.filter(sale__isnull=False, is_active=True).distinct()
        if is_admin else None
    )

    week_start  = today - timedelta(days=6)
    daily_sales = (
        Sale.objects.filter(date__gte=week_start)
        .annotate(day=ExtractDay('date'))
        .values('day')
        .annotate(total=Sum('total_amount'))
        .order_by('day')
    )
    daily_labels = [(week_start + timedelta(days=i)).strftime('%a') for i in range(7)]
    daily_data   = [0] * 7
    for ds in daily_sales:
        idx = (ds['day'] - week_start.day) % 7
        if 0 <= idx < 7:
            daily_data[idx] = float(ds['total'] or 0)

    context = {
        'products':              products,
        'sales':                 sale_query,
        'total_paid':            total_paid,
        'from_date':             from_date.strftime('%d-%m-%Y') if from_date else '',
        'to_date':               to_date.strftime('%d-%m-%Y')   if to_date   else '',
        'selected_product':      product_id,
        'period':                period,
        'chart_period':          chart_period,
        'total_to_be_paid':      total_to_be_paid,
        'remaining_balance':     remaining_balance,
        'total_items':           total_items,
        'customer_count':        customer_count,
        'sale_count':            sale_count,
        'quotation_count':       quotation_count,
        'registered_product':    registered_product,
        'this_month_sales':      this_month_sales,
        'last_month_sales':      last_month_sales,
        'sales_growth':          sales_growth,
        'chart_labels':          json.dumps(chart_labels),
        'chart_data':            json.dumps(chart_data),
        'chart_title':           chart_title,
        'monthly_totals':        monthly_totals,
        'monthly_labels':        monthly_labels,
        'months':                months,
        'pending_amounts':       pending_amounts,
        'partial_amounts':       partial_amounts,
        'full_amounts':          full_amounts,
        'product_chart_labels':  json.dumps(product_chart_labels),
        'product_chart_data':    json.dumps(product_chart_data),
        'sale_items_tot':        processed_items,
        'top_customers':         top_customers_qs,
        'daily_labels':          json.dumps(daily_labels),
        'daily_data':            json.dumps(daily_data),
        'is_admin':              is_admin,
        'users':                 users_with_sales,
        'selected_user_id':      selected_user_id,
    }
    return render(request, 'home/dash.html', context)


# ─────────────────────────────────────────────────────────────────────────────
# REPORTS
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def sales_report(request):
    from_date = request.GET.get('from_date')
    to_date   = request.GET.get('to_date')
    try:
        from_date = datetime.strptime(from_date, '%d-%m-%Y') if from_date else None
        to_date   = datetime.strptime(to_date,   '%d-%m-%Y') if to_date   else None
    except ValueError:
        from_date = to_date = None

    today     = timezone.now()
    yesterday = today - timedelta(days=1)

    is_admin = request.user.groups.filter(name__in=['Administrators', 'Salers']).exists()
    selected_user_id = request.GET.get('user_id') if is_admin else None

    sales_query = Sale.objects.all() if is_admin else Sale.objects.filter(creator=request.user)
    if is_admin and selected_user_id:
        sales_query = sales_query.filter(creator_id=selected_user_id)

    paid_sales_yesterday  = sales_query.filter(date__date=yesterday).aggregate(total=Sum('payments__paid_amount'))['total'] or 0
    paid_sales_today      = sales_query.filter(date__date=today).aggregate(total=Sum('payments__paid_amount'))['total'] or 0
    total_expected_amount = sales_query.filter(date__date=today).aggregate(total=Sum('total_amount'))['total'] or 0
    balance_sales         = total_expected_amount - paid_sales_today

    sale_items_query = SaleItem.objects.filter(sale__in=sales_query)
    if from_date:
        sale_items_query = sale_items_query.filter(sale__date__gte=from_date)
    if to_date:
        sale_items_query = sale_items_query.filter(sale__date__lte=to_date)

    sales_data = (
        sale_items_query
        .values('product_id', 'product__product_name', 'product__image',
                'sale__date', 'sale__creator__username', 'sale__customer__customer_name')
        .annotate(total_quantity=Sum('quantity'), total_sales=Sum('subtotal'))
        .order_by('-sale__date')
    )

    processed_items = [{
        'id':           item['product_id'],
        'name':         item['product__product_name'],
        'sale_date':    item['sale__date'].strftime('%Y-%m-%d') if item['sale__date'] else '',
        'quantity':     item['total_quantity'],
        'sold_total':   "{:,.2f}".format(item['total_sales']) if item['total_sales'] else "0.00",
        'image_url':    f"{settings.MEDIA_URL}{item['product__image']}" if item['product__image'] else f"{settings.STATIC_URL}img/noimage.png",
        'customer_name':item['sale__customer__customer_name'],
        'username':     item['sale__creator__username'],
    } for item in sales_data]

    context = {
        'sale_items_tot':       processed_items,
        'from_date':            from_date.strftime('%Y-%m-%d') if from_date else '',
        'to_date':              to_date.strftime('%Y-%m-%d')   if to_date   else '',
        'total_expected_amount':total_expected_amount,
        'paid_sales_today':     paid_sales_today,
        'balance_sales':        balance_sales,
        'paid_sales_yesterday': paid_sales_yesterday,
    }
    return render(request, 'home/report/salesreport.html', context)


def sales_report_pdf(request):
    from_date = request.GET.get('from_date')
    to_date   = request.GET.get('to_date')
    try:
        from_date = datetime.strptime(from_date, '%d-%m-%Y') if from_date else None
        to_date   = datetime.strptime(to_date,   '%d-%m-%Y') if to_date   else None
    except ValueError:
        from_date = to_date = None

    is_admin = request.user.groups.filter(name='Administrators').exists()
    selected_user_id = request.GET.get('user_id') if is_admin else None
    sales_query = Sale.objects.all() if is_admin else Sale.objects.filter(creator=request.user)
    if is_admin and selected_user_id:
        sales_query = sales_query.filter(creator_id=selected_user_id)

    sale_items_query = SaleItem.objects.filter(sale__in=sales_query)
    if from_date:
        sale_items_query = sale_items_query.filter(sale__date__gte=from_date)
    if to_date:
        sale_items_query = sale_items_query.filter(sale__date__lte=to_date)

    sale_items_tot = sale_items_query.values(
        'product__product_name', 'product__image', 'quantity', 'subtotal',
        'sale__date', 'sale__customer__customer_name', 'sale__creator__username'
    )
    total_quantity = sum(item['quantity'] for item in sale_items_tot)
    total_amount   = sum(item['subtotal'] for item in sale_items_tot)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'
    doc      = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles   = getSampleStyleSheet()

    summary_data  = [['Summary', ''], ['Total Quantity:', str(total_quantity)], ['Total Amount:', f"{total_amount:,} Rwf"]]
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0),(-1,0), colors.grey),
        ('TEXTCOLOR',  (0,0),(-1,0), colors.whitesmoke),
        ('SPAN',       (0,0),(1,0)),
        ('ALIGN',      (0,0),(-1,-1), 'LEFT'),
        ('FONTNAME',   (0,0),(-1,0), 'Helvetica-Bold'),
        ('FONTNAME',   (0,1),(0,-1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0),(-1,-1), 12),
        ('BOTTOMPADDING',(0,0),(-1,-1), 6),
        ('TOPPADDING', (0,0),(-1,-1), 6),
        ('GRID',       (0,0),(-1,-1), 1, colors.black),
        ('BACKGROUND', (0,1),(-1,-1), colors.beige),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    data = [['Product', 'Quantity', 'Total', 'Date', 'Customer', 'Username']]
    for item in sale_items_tot:
        data.append([
            item['product__product_name'], item['quantity'],
            f"{item['subtotal']:,} Rwf", item['sale__date'],
            item['sale__customer__customer_name'], item['sale__creator__username'],
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND',   (0,0),(-1,0), colors.grey),
        ('TEXTCOLOR',    (0,0),(-1,0), colors.whitesmoke),
        ('ALIGN',        (0,0),(-1,-1), 'CENTER'),
        ('FONTNAME',     (0,0),(-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0,0),(-1,0), 14),
        ('BOTTOMPADDING',(0,0),(-1,0), 12),
        ('BACKGROUND',   (0,1),(-1,-1), colors.beige),
        ('GRID',         (0,0),(-1,-1), 1, colors.black),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


@login_required
def sales_report_excel(request):
    from_date = request.GET.get('from_date')
    to_date   = request.GET.get('to_date')
    try:
        from_date = datetime.strptime(from_date, '%d-%m-%Y') if from_date else None
        to_date   = datetime.strptime(to_date,   '%d-%m-%Y') if to_date   else None
    except ValueError:
        from_date = to_date = None

    is_admin = request.user.groups.filter(name='Administrators').exists()
    selected_user_id = request.GET.get('user_id') if is_admin else None
    sales_query = Sale.objects.all() if is_admin else Sale.objects.filter(creator=request.user)
    if is_admin and selected_user_id:
        sales_query = sales_query.filter(creator_id=selected_user_id)

    sale_items_query = SaleItem.objects.filter(sale__in=sales_query)
    if from_date:
        sale_items_query = sale_items_query.filter(sale__date__gte=from_date)
    if to_date:
        sale_items_query = sale_items_query.filter(sale__date__lte=to_date)

    sale_items_tot = sale_items_query.values(
        'product__product_name', 'quantity', 'subtotal',
        'sale__date', 'sale__customer__customer_name', 'sale__creator__username'
    )

    workbook  = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Sales Report'

    header_fill = PatternFill(start_color='1E2D40', end_color='1E2D40', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    headers     = ['Product', 'Quantity', 'Total (Rwf)', 'Date', 'Customer', 'Username']
    for col, h in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    row = 2
    for item in sale_items_tot:
        worksheet.cell(row=row, column=1, value=item['product__product_name'])
        worksheet.cell(row=row, column=2, value=item['quantity'])
        worksheet.cell(row=row, column=3, value=float(item['subtotal']))
        worksheet.cell(row=row, column=4, value=item['sale__date'].replace(tzinfo=None))
        worksheet.cell(row=row, column=5, value=item['sale__customer__customer_name'])
        worksheet.cell(row=row, column=6, value=item['sale__creator__username'])
        row += 1

    for col in worksheet.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        worksheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sales_report.xlsx'
    workbook.save(response)
    return response


@login_required
def top_customers(request):
    products   = Product.objects.values('productid', 'product_name').order_by('product_name')
    product_id = request.GET.get('product_id', '')
    try:
        product_id = int(product_id) if product_id else None
    except ValueError:
        product_id = None

    customers_query = SaleItem.objects.values(
        'sale__customer__customer_name',
        'sale__customer__cust_id',
    ).annotate(
        total_items =Coalesce(Sum('quantity'), Value(0), output_field=IntegerField()),
        total_amount=Coalesce(Sum('subtotal'), Value(0), output_field=DecimalField()),
    ).order_by('-total_amount')

    if product_id:
        customers_query = customers_query.filter(product__productid=product_id)

    context = {
        'top_customers':       customers_query[:10],
        'products':            products,
        'selected_product_id': product_id,
    }
    return render(request, 'home/report/customer_report1.html', context)


@login_required
def low_stock_alert(request):
    """
    Returns products with available stock <= 10.
    Uses same pandas logic as productlist view:
    available = total stock added - total sold quantity
    """
    LOW_STOCK_THRESHOLD = 10
 
    # ── Replicate exact productlist logic ─────────────────────────────────
    from product.models import Stock as StockModel
 
    stock_df    = pd.DataFrame(list(StockModel.objects.all().values('product_id', 'stock_quantity')))
    sale_df     = pd.DataFrame(list(SaleItem.objects.all().values('product_id', 'quantity')))
    products_df = pd.DataFrame(list(
        Product.objects.all().values('productid', 'product_name', 'brand', 'image')
    ))
 
    if products_df.empty:
        return JsonResponse({'count': 0, 'items': [], 'threshold': LOW_STOCK_THRESHOLD})
 
    # Aggregate stock and sales
    if stock_df.empty:
        stock_qty = pd.DataFrame(columns=['product_id', 'stock_quantity'])
    else:
        stock_qty = stock_df.groupby('product_id')['stock_quantity'].sum().reset_index()
 
    if sale_df.empty:
        sale_qty = pd.DataFrame(columns=['product_id', 'quantity'])
    else:
        sale_qty = sale_df.groupby('product_id')['quantity'].sum().reset_index()
 
    # Merge and calculate available stock
    stock_balance = stock_qty.merge(sale_qty, on='product_id', how='outer')
    stock_balance['stock_quantity'] = stock_balance['stock_quantity'].fillna(0)
    stock_balance['quantity']       = stock_balance['quantity'].fillna(0)
    stock_balance['available']      = stock_balance['stock_quantity'] - stock_balance['quantity']
 
    # Merge with products
    merged = products_df.merge(
        stock_balance[['product_id', 'stock_quantity', 'available']],
        left_on='productid',
        right_on='product_id',
        how='left'
    )
    merged['available']      = merged['available'].fillna(0)
    merged['stock_quantity'] = merged['stock_quantity'].fillna(0)
 
    # Filter to low stock only, order by most critical first
    low = merged[merged['available'] <= LOW_STOCK_THRESHOLD].sort_values('available')
 
    items = []
    for _, row in low.iterrows():
        image_url = ''
        if pd.notna(row['image']) and str(row['image']) != '':
            image_url = settings.MEDIA_URL + str(row['image'])
        items.append({
            'id':          int(row['productid']),
            'name':        str(row['product_name']),
            'brand':       str(row['brand']) if pd.notna(row['brand']) else '—',
            'available':   int(row['available']),
            'total_stock': int(row['stock_quantity']),
            'image':       image_url,
        })
 
    return JsonResponse({
        'count':     len(items),
        'items':     items,
        'threshold': LOW_STOCK_THRESHOLD,
    })